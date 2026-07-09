import json
import math

import pandas as pd

from django.apps import apps
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import TemplateView

from .models import CargaArchivo, FuenteDatos, MapeoColumna, Proyecto, Reportador, Sitio


class DashboardView(TemplateView):
    template_name = "app/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["sitio_count"] = Sitio.objects.count()
        context["proyecto_count"] = Proyecto.objects.count()
        return context


def fuentes_datos_api(request):
    proyecto_id = request.GET.get("proyecto")
    qs = FuenteDatos.objects.select_related("proyecto", "reportador")
    if proyecto_id:
        qs = qs.filter(proyecto_id=proyecto_id)

    fuentes = []
    for f in qs:
        fuentes.append({
            "id": f.id,
            "nombre": f.nombre,
            "descripcion": f.descripcion,
            "tipo": f.tipo,
            "tipo_label": f.get_tipo_display(),
            "url": f.url,
            "estado": f.estado,
            "estado_label": f.get_estado_display(),
            "responsable": f.reportador.nombre if f.reportador else "",
            "fecha_datos": f.fecha_recepcion.isoformat() if f.fecha_recepcion else None,
            "notas": f.notas,
            "proyecto": {"id": f.proyecto.id, "codigo": f.proyecto.codigo, "nombre": f.proyecto.nombre} if f.proyecto else None,
            "sitio": None,
            "created_at": f.created_at.isoformat(),
        })

    proyectos = []
    for p in Proyecto.objects.all():
        proyectos.append({"id": p.id, "codigo": p.codigo, "nombre": p.nombre})

    return JsonResponse({"fuentes": fuentes, "proyectos": proyectos}, json_dumps_params={"ensure_ascii": False})


@csrf_exempt
def crear_fuente_datos(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)

    nombre = (body.get("nombre") or "").strip()
    if not nombre:
        return JsonResponse({"error": "El nombre es obligatorio"}, status=400)

    proyecto_id = body.get("proyecto_id") or None
    proyecto = None
    if proyecto_id:
        try:
            proyecto = Proyecto.objects.get(pk=proyecto_id)
        except Proyecto.DoesNotExist:
            return JsonResponse({"error": "Proyecto no encontrado"}, status=404)

    responsable = (body.get("responsable") or "").strip()
    reportador = None
    if responsable:
        reportador, _ = Reportador.objects.get_or_create(nombre=responsable)

    fuente = FuenteDatos.objects.create(
        nombre=nombre,
        descripcion=body.get("descripcion", ""),
        url=body.get("url", ""),
        tipo=body.get("tipo", "excel"),
        estado=body.get("estado", "pendiente"),
        reportador=reportador,
        proyecto=proyecto,
    )

    return JsonResponse({
        "id": fuente.id,
        "nombre": fuente.nombre,
        "tipo": fuente.tipo,
        "tipo_label": fuente.get_tipo_display(),
        "estado": fuente.estado,
        "estado_label": fuente.get_estado_display(),
        "url": fuente.url,
        "descripcion": fuente.descripcion,
        "responsable": fuente.reportador.nombre if fuente.reportador else "",
        "proyecto": {"id": proyecto.id, "codigo": proyecto.codigo, "nombre": proyecto.nombre} if proyecto else None,
        "sitio": None,
        "fecha_datos": None,
        "notas": "",
        "created_at": fuente.created_at.isoformat(),
    }, status=201, json_dumps_params={"ensure_ascii": False})


@csrf_exempt
def crear_proyecto(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)

    codigo = (body.get("codigo") or "").strip()
    nombre = (body.get("nombre") or "").strip()
    if not codigo:
        return JsonResponse({"error": "El código es obligatorio"}, status=400)
    if not nombre:
        return JsonResponse({"error": "El nombre es obligatorio"}, status=400)
    if Proyecto.objects.filter(codigo=codigo).exists():
        return JsonResponse({"error": f"Ya existe un proyecto con el código «{codigo}»"}, status=400)

    proyecto = Proyecto.objects.create(
        codigo=codigo,
        nombre=nombre,
        coordinador=body.get("coordinador", ""),
        correo_coordinador=body.get("correo_coordinador", ""),
        escala_espacial=body.get("escala_espacial", ""),
        objetivo_general=body.get("objetivo_general", ""),
        fecha_inicio=body.get("fecha_inicio") or None,
        fecha_fin=body.get("fecha_fin") or None,
    )

    return JsonResponse({
        "id": proyecto.id,
        "codigo": proyecto.codigo,
        "nombre": proyecto.nombre,
        "coordinador": proyecto.coordinador,
        "correo_coordinador": proyecto.correo_coordinador,
        "escala_espacial": proyecto.escala_espacial,
        "objetivo_general": proyecto.objetivo_general,
        "fecha_inicio": proyecto.fecha_inicio.isoformat() if proyecto.fecha_inicio else None,
        "fecha_fin": proyecto.fecha_fin.isoformat() if proyecto.fecha_fin else None,
    }, status=201, json_dumps_params={"ensure_ascii": False})


def _infer_dtype(series):
    if pd.api.types.is_bool_dtype(series):
        return "boolean"
    if pd.api.types.is_numeric_dtype(series):
        return "number"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "date"
    return "string"


def _muestra(series, n=3):
    valores = series.dropna().head(n).tolist()
    return [str(v) for v in valores]


@csrf_exempt
def upload_archivo(request, fuente_id):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        fuente = FuenteDatos.objects.get(pk=fuente_id)
    except FuenteDatos.DoesNotExist:
        return JsonResponse({"error": "Fuente de datos no encontrada"}, status=404)

    archivo = request.FILES.get("archivo")
    if not archivo:
        return JsonResponse({"error": "No se recibió ningún archivo"}, status=400)

    nombre = archivo.name.lower()
    if not (nombre.endswith(".xlsx") or nombre.endswith(".xls") or nombre.endswith(".csv")):
        return JsonResponse({"error": "Formato no permitido. Solo .xlsx, .xls o .csv"}, status=400)

    try:
        carga = CargaArchivo.objects.create(fuente=fuente, archivo=archivo)

        if nombre.endswith(".csv"):
            df = pd.read_csv(carga.archivo.path)
            sheets = []
            hoja_activa = ""
            total_filas = len(df)
            columnas = []
            for col in df.columns:
                columnas.append({
                    "nombre": col,
                    "dtype": _infer_dtype(df[col]),
                    "nulls": int(df[col].isna().sum()),
                    "muestra": _muestra(df[col]),
                })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(carga.archivo.path, read_only=True, data_only=True)
            sheets = wb.sheetnames

            hoja_activa = sheets[0]
            for sheet_name in sheets:
                ws = wb[sheet_name]
                if ws.max_row and ws.max_row > 1:
                    hoja_activa = sheet_name
                    break

            df = pd.read_excel(carga.archivo.path, sheet_name=hoja_activa)
            total_filas = len(df)
            columnas = []
            for col in df.columns:
                columnas.append({
                    "nombre": col,
                    "dtype": _infer_dtype(df[col]),
                    "nulls": int(df[col].isna().sum()),
                    "muestra": _muestra(df[col]),
                })
            wb.close()

        carga.hoja_activa = hoja_activa
        carga.columnas_raw = columnas
        carga.total_filas = total_filas
        carga.save(update_fields=["hoja_activa", "columnas_raw", "total_filas"])

        return JsonResponse({
            "carga_id": carga.pk,
            "sheets": sheets,
            "hoja_activa": hoja_activa,
            "total_filas": total_filas,
            "columnas": columnas,
        }, json_dumps_params={"ensure_ascii": False})

    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


def campos_destino(request):
    CAMPOS_DESTINO = {
        "Publicacion": ["titulo", "doi", "anio", "url", "resumen"],
        "Sitio": ["nombre", "latitud", "longitud", "altitud"],
        "Cobertura": ["nombre", "cobertura_ipcc", "cobertura_clc"],
        "Parcela": ["nombre", "area_ha"],
        "MonitoreoSuelo": ["fecha", "profundidad_cm"],
    }
    return JsonResponse({"modelos": CAMPOS_DESTINO}, json_dumps_params={"ensure_ascii": False})


@csrf_exempt
def mapeo_carga(request, fuente_id, carga_id):
    try:
        carga = CargaArchivo.objects.get(pk=carga_id, fuente_id=fuente_id)
    except CargaArchivo.DoesNotExist:
        return JsonResponse({"error": "Carga no encontrada"}, status=404)

    if request.method == "GET":
        mapeos = list(
            carga.mapeos.values("columna_origen", "modelo_destino", "campo_destino", "transformacion")
        )
        return JsonResponse({"carga_id": carga.pk, "mapeos": mapeos}, json_dumps_params={"ensure_ascii": False})

    if request.method == "POST":
        try:
            body = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "JSON inválido"}, status=400)

        items = body.get("mapeos")
        if not isinstance(items, list):
            return JsonResponse({"error": "Se esperaba un array en 'mapeos'"}, status=400)

        try:
            guardados = 0
            for item in items:
                columna_origen = (item.get("columna_origen") or "").strip()
                if not columna_origen:
                    continue
                MapeoColumna.objects.update_or_create(
                    carga=carga,
                    columna_origen=columna_origen,
                    defaults={
                        "modelo_destino": item.get("modelo_destino", ""),
                        "campo_destino": item.get("campo_destino", ""),
                        "transformacion": item.get("transformacion", "directo"),
                    },
                )
                guardados += 1

            carga.estado = "mapeado"
            carga.save(update_fields=["estado"])

            return JsonResponse({"ok": True, "guardados": guardados})
        except Exception as exc:
            return JsonResponse({"error": str(exc)}, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)


GRUPOS_CATALOGO = [
    {
        "nombre": "Publicaciones",
        "icono": "📄",
        "entidades": ["PublicacionType", "Publicacion", "Autor"],
    },
    {
        "nombre": "Geografía",
        "icono": "🗺️",
        "entidades": ["Region", "Departamento", "Municipio"],
    },
    {
        "nombre": "Sitio y Parcelas",
        "icono": "📍",
        "entidades": ["Sitio", "Parcela", "Transecto"],
    },
    {
        "nombre": "Cobertura y Vegetación",
        "icono": "🌿",
        "entidades": ["Cobertura", "Vegetacion", "Disturbio"],
    },
    {
        "nombre": "Suelo",
        "icono": "🪨",
        "entidades": ["CaracterizacionMuestreoSuelo", "MonitoreoSuelo"],
    },
    {
        "nombre": "Torre EC y Flujos",
        "icono": "📡",
        "entidades": ["TorreEc", "Equipo", "ConfiguracionSensorGas", "FlujoCamaras"],
    },
    {
        "nombre": "Proyecto",
        "icono": "🗂️",
        "entidades": ["Proyecto", "Aliado"],
    },
]

ENTIDADES_SEMILLA = {
    "PublicacionType",
}

TIPO_MAP = {
    'CharField': 'Texto',
    'TextField': 'Texto largo',
    'IntegerField': 'Entero',
    'FloatField': 'Decimal',
    'DecimalField': 'Decimal',
    'DateField': 'Fecha',
    'DateTimeField': 'Fecha y hora',
    'BooleanField': 'Booleano',
    'URLField': 'URL',
    'EmailField': 'Correo',
    'JSONField': 'JSON',
    'FileField': 'Archivo',
    'ForeignKey': 'FK (relación)',
}


def _semilla_rows(modelo_cls):
    concrete_fields = [
        f.attname for f in modelo_cls._meta.get_fields()
        if hasattr(f, 'column')
    ]
    return list(modelo_cls.objects.values(*concrete_fields))


@csrf_exempt
def catalogo_tecnico(request):
    grupos = []
    for grupo_def in GRUPOS_CATALOGO:
        entidades = []
        for nombre_entidad in grupo_def["entidades"]:
            try:
                modelo_cls = apps.get_model('app', nombre_entidad)
            except LookupError:
                continue

            campos = []
            for field in modelo_cls._meta.get_fields():
                if field.is_relation and not hasattr(field, 'column'):
                    continue
                if field.name in ('id', 'created_at', 'updated_at'):
                    continue

                tipo_raw = field.__class__.__name__
                tipo_legible = TIPO_MAP.get(tipo_raw, tipo_raw)

                campos.append({
                    'nombre': field.name,
                    'verbose_name': str(getattr(field, 'verbose_name', field.name)),
                    'tipo': tipo_legible,
                    'tipo_raw': tipo_raw,
                    'requerido': not (getattr(field, 'blank', True) or getattr(field, 'null', True)),
                    'max_length': getattr(field, 'max_length', None),
                    'choices': [{'valor': c[0], 'etiqueta': c[1]} for c in getattr(field, 'choices', []) or []],
                    'es_fk': tipo_raw == 'ForeignKey',
                    'modelo_fk': field.related_model.__name__ if tipo_raw == 'ForeignKey' else None,
                })

            entry = {
                'nombre': nombre_entidad,
                'verbose_name': str(modelo_cls._meta.verbose_name),
                'verbose_name_plural': str(modelo_cls._meta.verbose_name_plural),
                'total_campos': len(campos),
                'campos': campos,
            }
            if nombre_entidad in ENTIDADES_SEMILLA:
                entry['datos_semilla'] = _semilla_rows(modelo_cls)
            entidades.append(entry)

        grupos.append({
            'nombre': grupo_def['nombre'],
            'icono': grupo_def['icono'],
            'entidades': entidades,
        })

    return JsonResponse({'grupos': grupos}, json_dumps_params={'ensure_ascii': False})


MODELOS_ERD = [
    "Region", "Departamento", "Municipio",
    "Cobertura", "Vegetacion", "Disturbio",
    "Sitio", "Parcela", "Transecto", "MonitoreoParcela",
    "Proyecto", "Aliado",
    "PublicacionType", "Publicacion", "Autor", "PublicacionSitio", "ResultadoPublicacion",
    "CaracterizacionMuestreoSuelo", "MonitoreoSuelo",
    "TorreEc", "Equipo", "ConfiguracionSensorGas", "FlujoCamaras",
    "FuenteDatos", "Reportador",
]

TIPO_MAP_ERD = {
    'CharField': 'string', 'TextField': 'string', 'URLField': 'string',
    'EmailField': 'string', 'SlugField': 'string',
    'IntegerField': 'int', 'SmallIntegerField': 'int', 'PositiveIntegerField': 'int',
    'FloatField': 'float', 'DecimalField': 'decimal',
    'DateField': 'date', 'DateTimeField': 'datetime',
    'BooleanField': 'boolean',
    'JSONField': 'json',
    'FileField': 'string',
    'ForeignKey': 'int',
}


def generar_mermaid_erd():
    lines = ["erDiagram"]
    relaciones = []

    for nombre_modelo in MODELOS_ERD:
        try:
            modelo_cls = apps.get_model('app', nombre_modelo)
        except LookupError:
            continue

        campos_erd = ['    %s {' % nombre_modelo]

        for field in modelo_cls._meta.get_fields():
            if field.is_relation and not hasattr(field, 'column'):
                continue
            if field.name in ('id', 'created_at', 'updated_at'):
                continue

            tipo_raw = field.__class__.__name__
            tipo_erd = TIPO_MAP_ERD.get(tipo_raw, 'string')
            nombre_campo = field.name

            if tipo_raw == 'ForeignKey':
                nombre_campo = field.name + '_id'
                modelo_destino = field.related_model.__name__
                if modelo_destino in MODELOS_ERD:
                    card_izq = '}o' if getattr(field, 'null', False) else '}|'
                    card_der = 'o|' if getattr(field, 'null', False) else '||'
                    relaciones.append(
                        f'    {nombre_modelo} {card_izq}--{card_der} {modelo_destino} : "{field.name}"'
                    )

            campos_erd.append(f'        {tipo_erd} {nombre_campo}')

        campos_erd.append('    }')
        lines.extend(campos_erd)

    lines.extend(relaciones)
    return '\n'.join(lines)


@csrf_exempt
def db_erd(request):
    try:
        mermaid_str = generar_mermaid_erd()
        return JsonResponse({
            'mermaid': mermaid_str,
            'total_modelos': len(MODELOS_ERD),
            'generado_en': __import__('datetime').datetime.now().isoformat(),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def _to_python(val):
    if val is None:
        return None
    if isinstance(val, float) and math.isnan(val):
        return None
    try:
        import numpy as np
        if isinstance(val, (np.integer,)):
            return int(val)
        if isinstance(val, (np.floating,)):
            return None if math.isnan(float(val)) else float(val)
        if isinstance(val, (np.bool_,)):
            return bool(val)
    except ImportError:
        pass
    return val


def _es_vacio(val):
    if val is None:
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    if isinstance(val, str) and val.strip() == "":
        return True
    return False


def _validar_columna(df, columna_origen, modelo_destino, campo_destino):
    try:
        modelo_cls = apps.get_model("app", modelo_destino)
    except LookupError:
        return {"advertencia": "modelo_o_campo_no_encontrado"}

    try:
        field = modelo_cls._meta.get_field(campo_destino)
    except Exception:
        return {"advertencia": "modelo_o_campo_no_encontrado"}

    if columna_origen not in df.columns:
        return {"advertencia": "modelo_o_campo_no_encontrado"}

    serie = df[columna_origen]
    errores = []

    campo_requerido = not getattr(field, "blank", True) and not getattr(field, "null", True)
    max_length = getattr(field, "max_length", None)
    choices = getattr(field, "choices", None)
    choices_valores = {c[0] for c in choices} if choices else None

    tipo_campo = type(field).__name__

    for idx, val in serie.items():
        fila = int(idx) + 2
        py_val = _to_python(val)

        if _es_vacio(val):
            if campo_requerido:
                errores.append({
                    "fila": fila,
                    "valor": None,
                    "tipo": "nulo_obligatorio",
                    "mensaje": "Campo requerido vacío",
                })
            continue

        if tipo_campo in ("FloatField", "DecimalField"):
            try:
                float(val)
            except (ValueError, TypeError):
                errores.append({
                    "fila": fila,
                    "valor": py_val,
                    "tipo": "tipo_invalido",
                    "mensaje": "No se puede convertir a número",
                })
                continue

        elif tipo_campo in ("IntegerField", "PositiveIntegerField", "PositiveSmallIntegerField",
                            "SmallIntegerField", "BigIntegerField"):
            try:
                int(float(val))
            except (ValueError, TypeError):
                errores.append({
                    "fila": fila,
                    "valor": py_val,
                    "tipo": "tipo_invalido",
                    "mensaje": "No se puede convertir a entero",
                })
                continue

        elif tipo_campo in ("DateField", "DateTimeField"):
            try:
                pd.to_datetime(val)
            except Exception:
                errores.append({
                    "fila": fila,
                    "valor": py_val,
                    "tipo": "tipo_invalido",
                    "mensaje": "No se puede interpretar como fecha",
                })
                continue

        if max_length is not None:
            if len(str(val)) > max_length:
                errores.append({
                    "fila": fila,
                    "valor": py_val,
                    "tipo": "longitud_excedida",
                    "mensaje": f"Longitud {len(str(val))} supera el máximo de {max_length}",
                })
                continue

        if choices_valores is not None:
            if val not in choices_valores and str(val) not in choices_valores:
                errores.append({
                    "fila": fila,
                    "valor": py_val,
                    "tipo": "fuera_de_vocabulario",
                    "mensaje": f"Valor no permitido. Opciones: {sorted(str(v) for v in choices_valores)}",
                })

    total = len(serie)
    return {
        "columna": columna_origen,
        "modelo_destino": modelo_destino,
        "campo_destino": campo_destino,
        "total": total,
        "ok": total - len(errores),
        "errores": errores,
    }


@csrf_exempt
def validar_carga(request, fuente_id, carga_id):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        carga = CargaArchivo.objects.get(pk=carga_id, fuente_id=fuente_id)
    except CargaArchivo.DoesNotExist:
        return JsonResponse({"error": "Carga no encontrada"}, status=404)

    try:
        path = carga.archivo.path
        nombre = path.lower()
        if nombre.endswith(".csv"):
            df = pd.read_csv(path)
        else:
            df = pd.read_excel(path, sheet_name=carga.hoja_activa)

        mapeos = carga.mapeos.exclude(modelo_destino="")

        resultados = []
        filas_con_error = set()

        for mapeo in mapeos:
            resultado = _validar_columna(df, mapeo.columna_origen, mapeo.modelo_destino, mapeo.campo_destino)
            if "advertencia" not in resultado:
                for e in resultado["errores"]:
                    filas_con_error.add(e["fila"])
            resultados.append(resultado)

        columnas_con_errores = sum(
            1 for r in resultados
            if "advertencia" not in r and len(r["errores"]) > 0
        )
        total_errores = sum(
            len(r["errores"]) for r in resultados if "advertencia" not in r
        )

        return JsonResponse({
            "resumen": {
                "total_filas": carga.total_filas,
                "columnas_mapeadas": mapeos.count(),
                "columnas_con_errores": columnas_con_errores,
                "total_errores": total_errores,
                "filas_limpias": carga.total_filas - len(filas_con_error),
                "filas_con_errores": len(filas_con_error),
            },
            "columnas": resultados,
        }, json_dumps_params={"ensure_ascii": False})

    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)
