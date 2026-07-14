import csv
import json
import math
import re
import urllib.request
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd

from django.apps import apps
from django.conf import settings
from django.db import transaction
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

from app.models import CargaArchivo, FuenteDatos, MapeoColumna

from app.catalogo.generator import GRUPOS_CATALOGO, campo_to_catalogo, fk_choices

# El wizard del ETL usa su propio agrupamiento de secciones, más fino que
# GRUPOS_CATALOGO (que sigue usándose tal cual para el catálogo de
# referencia en docs/):
# - Separa "Unidad Experimental" de "Unidad de Muestreo" porque en los
#   archivos reales primero se identifica la unidad experimental y solo
#   después se resuelve el tipo/unidad de muestreo/parcela que dependen de
#   ella (vía la FK a UnidadExperimental mapeada en la sección anterior).
# - Excluye "Geografía" (Region/Departamento/Municipio) salvo "Sitio", que
#   se saca de ese grupo y se deja como su propia sección al final: en el
#   catálogo vive junto a la geografía porque se georreferencia por
#   Municipio, pero en el wizard depende de que ya exista la Unidad de
#   Muestreo (no siempre se completa con datos del archivo).
# - Excluye "Publicaciones": es metadato bibliográfico del proyecto, no
#   datos fila por fila del archivo. Al quitarla, "Unidad Experimental"
#   queda como primera sección real del wizard.
# - Excluye temporalmente "Cobertura y Vegetación", "Suelo", "Torre EC y
#   Flujos", "Proyecto" y "Usuarios, Roles y ETL": todavía no se está
#   mapeando/importando esas entidades desde el ETL. Por ahora el wizard
#   llega hasta "Muestras CO₂". Quitar de esta lista cuando se retome cada
#   una.
_GRUPOS_EXCLUIDOS_TEMPORAL = (
    "Cobertura y Vegetación",
    "Suelo",
    "Torre EC y Flujos",
    "Proyecto",
    "Usuarios, Roles y ETL",
)
SECCIONES_ETL = []
for _grupo in GRUPOS_CATALOGO:
    if _grupo["nombre"] in ("Geografía", "Publicaciones") or _grupo["nombre"] in _GRUPOS_EXCLUIDOS_TEMPORAL:
        continue
    if _grupo["nombre"] == "Unidad de Muestreo y Experimental":
        SECCIONES_ETL.append({
            "nombre": "Unidad Experimental",
            "icono": "🧪",
            "entidades": ["UnidadExperimental"],
        })
        SECCIONES_ETL.append({
            "nombre": "Unidad de Muestreo",
            "icono": "📏",
            # UnidadMuestreoTipo no se incluye aquí a propósito: es un catálogo
            # cerrado de tipos (parcela, transecto, etc.), no datos que el ETL
            # deba crear o modificar. El campo "tipo" de UnidadMuestreo se sigue
            # pudiendo mapear igual, pero como FK de solo selección entre los
            # tipos ya existentes (ver fk_choices en catalogo/generator.py).
            "entidades": ["UnidadMuestreo", "Parcela", "Transecto"],
        })
        SECCIONES_ETL.append({
            "nombre": "Sitio",
            "icono": "📍",
            "entidades": ["Sitio"],
        })
    else:
        SECCIONES_ETL.append(_grupo)


def _infer_dtype(series):
    if pd.api.types.is_bool_dtype(series):
        return "boolean"
    if pd.api.types.is_numeric_dtype(series):
        return "number"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "date"
    return "string"


def _muestra(series, n=3):
    valores = series.dropna().head(n).tolist()
    return [str(v) for v in valores]


def _sugerencias_hora(valores_unicos):
    """Para columnas que puedan mapearse a un TimeField: de sus valores únicos,
    cuáles tienen forma de hora ambigua (24h + sufijo a. m./p. m. contradictorio)
    y qué interpretación se les podría sugerir al usuario para que la confirme
    o corrija en la UI (nunca se aplica sola)."""
    sugerencias = {}
    for val in valores_unicos:
        sugerencia = _sugerir_hora(val)
        if sugerencia is not None:
            sugerencias[val] = sugerencia
    return sugerencias


_RE_HORA_LAXA = re.compile(r"^\d{1,2}:\d{2}(:\d{2}(\.\d+)?)?\s*([ap]\.?\s*m\.?)?\s*$", re.IGNORECASE)


def _interpretaciones_hora(valores_unicos):
    """Para toda columna con pinta de hora: cómo se interpretaría cada valor
    único en 24h (HH:MM:ss) tal como quedaría guardado en la base de datos.
    Se usa para mostrarle al usuario, antes de guardar, exactamente en qué se
    convierte cada valor de origen -incluidos los que ya son válidos y no
    requieren ninguna corrección- y así detectar ambigüedades a simple vista."""
    interpretaciones = {}
    for val in valores_unicos:
        if not _RE_HORA_LAXA.match(str(val).strip()):
            continue
        try:
            interpretaciones[val] = pd.to_datetime(val).strftime("%H:%M:%S")
            continue
        except Exception:
            pass
        sugerencia = _sugerir_hora(val)
        if sugerencia is not None:
            interpretaciones[val] = sugerencia
    return interpretaciones


def _columnas_desde_dataframe(df):
    columnas = []
    for col in df.columns:
        # Los valores únicos para el panel de choices se limitan a los
        # primeros max_n (rendimiento), pero la detección de horas ambiguas
        # necesita revisar TODOS los valores: un valor problemático puede
        # aparecer más adelante en la columna y quedar fuera de ese límite.
        todos_unicos = [str(v) for v in df[col].dropna().unique()]
        columnas.append({
            "nombre": col,
            "dtype": _infer_dtype(df[col]),
            "nulls": int(df[col].isna().sum()),
            "muestra": _muestra(df[col]),
            "valores_unicos": todos_unicos[:50],
            "sugerencias_hora": _sugerencias_hora(todos_unicos),
            "interpretaciones_hora": _interpretaciones_hora(todos_unicos),
        })
    return columnas


_EXTENSIONES_VALIDAS = (".xlsx", ".xls", ".csv")

_EXT_POR_CONTENT_TYPE = {
    "text/csv": ".csv",
    "application/csv": ".csv",
    "application/vnd.ms-excel": ".xls",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
}


def _descargar_archivo_remoto(fuente, url):
    cache_dir = Path(settings.MEDIA_ROOT) / "fuentes_datos_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)

    try:
        req = urllib.request.Request(url, headers={"User-Agent": "colflux-etl/1.0"})
        with urllib.request.urlopen(req, timeout=20) as resp:
            content_type = resp.headers.get_content_type()
            data = resp.read()
    except Exception as exc:
        raise ValueError(f"No se pudo descargar el archivo remoto: {exc}") from exc

    ext = Path(urlparse(url).path).suffix.lower()
    if ext not in _EXTENSIONES_VALIDAS:
        ext = _EXT_POR_CONTENT_TYPE.get(content_type, "")
    if ext not in _EXTENSIONES_VALIDAS:
        raise ValueError(
            "No se pudo determinar el formato del archivo remoto (.xlsx, .xls o .csv)."
        )

    destino = cache_dir / f"fuente_{fuente.pk}{ext}"
    destino.write_bytes(data)
    return destino


def _guardar_archivo_subido(fuente, archivo):
    nombre = archivo.name.lower()
    if not nombre.endswith(_EXTENSIONES_VALIDAS):
        raise ValueError("Formato no permitido. Solo .xlsx, .xls o .csv")

    destino_dir = Path(settings.MEDIA_ROOT) / "fuentes_datos"
    destino_dir.mkdir(parents=True, exist_ok=True)
    destino = destino_dir / f"fuente_{fuente.pk}{Path(nombre).suffix}"

    with open(destino, "wb") as f:
        for chunk in archivo.chunks():
            f.write(chunk)

    fuente.url = str(destino.relative_to(settings.BASE_DIR))
    fuente.save(update_fields=["url"])
    return destino


def _detectar_separador(muestra):
    lineas = [l for l in muestra.splitlines()[:20] if l.strip()]
    if not lineas:
        return ","

    mejor_sep, mejor_puntaje = ",", -1
    for sep in (",", ";", "\t", "|"):
        conteos = [len(next(csv.reader([l], delimiter=sep))) - 1 for l in lineas]
        if min(conteos) == 0:
            continue
        # Preferimos el separador que produce más columnas de forma consistente
        consistentes = sum(1 for c in conteos if c == conteos[0])
        puntaje = conteos[0] * consistentes
        if puntaje > mejor_puntaje:
            mejor_sep, mejor_puntaje = sep, puntaje
    return mejor_sep


def _leer_csv(path, **kwargs):
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with open(path, encoding=encoding, newline="") as f:
                muestra = f.read(64 * 1024)
        except UnicodeDecodeError:
            continue

        sep = _detectar_separador(muestra)
        try:
            return pd.read_csv(path, encoding=encoding, sep=sep, **kwargs)
        except UnicodeDecodeError:
            continue
        except pd.errors.ParserError as exc:
            raise ValueError(
                "El archivo CSV tiene filas con distinta cantidad de columnas "
                f"(separador detectado: '{sep}'). Revisa que no tenga filas de "
                "título antes del encabezado ni celdas con el separador sin "
                f"comillas. Detalle: {exc}"
            ) from exc
    raise ValueError("No se pudo determinar la codificación del archivo CSV.")


def _resolver_ruta_fuente(fuente):
    raw_path = (fuente.url or "").strip()
    if not raw_path:
        raise ValueError("La fuente no tiene enlace o ruta de archivo registrada.")

    parsed = urlparse(raw_path)
    if parsed.scheme in ("http", "https"):
        return _descargar_archivo_remoto(fuente, raw_path)
    if parsed.scheme == "file":
        raw_path = parsed.path
    elif parsed.scheme:
        raise ValueError("El enlace de la fuente usa un esquema no soportado.")

    path = Path(raw_path).expanduser()
    if not path.is_absolute():
        path = Path(settings.BASE_DIR) / path
    path = path.resolve()

    if not path.is_file():
        raise FileNotFoundError("No se encontró el archivo registrado en la fuente.")

    return path


@csrf_exempt
def archivo_fuente(request, fuente_id):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        fuente = FuenteDatos.objects.get(pk=fuente_id)
    except FuenteDatos.DoesNotExist:
        return JsonResponse({"error": "Fuente de datos no encontrada"}, status=404)

    archivo_subido = request.FILES.get("archivo")
    if not archivo_subido:
        return JsonResponse({"error": "No se recibió ningún archivo"}, status=400)

    try:
        _guardar_archivo_subido(fuente, archivo_subido)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    return JsonResponse({"url": fuente.url})


@csrf_exempt
def upload_archivo(request, fuente_id):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        fuente = FuenteDatos.objects.get(pk=fuente_id)
    except FuenteDatos.DoesNotExist:
        return JsonResponse({"error": "Fuente de datos no encontrada"}, status=404)

    archivo_subido = request.FILES.get("archivo")

    try:
        if archivo_subido:
            path_archivo = _guardar_archivo_subido(fuente, archivo_subido)
        else:
            path_archivo = _resolver_ruta_fuente(fuente)

        nombre_archivo = path_archivo.name.lower()
        if not nombre_archivo.endswith(_EXTENSIONES_VALIDAS):
            return JsonResponse({"error": "Formato no permitido. Solo .xlsx, .xls o .csv"}, status=400)

        carga = CargaArchivo.objects.create(fuente=fuente)

        if nombre_archivo.endswith(".csv"):
            df = _leer_csv(path_archivo)
            sheets = []
            hoja_activa = ""
            total_filas = len(df)
            columnas = _columnas_desde_dataframe(df)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(path_archivo, read_only=True, data_only=True)
            sheets = wb.sheetnames

            hoja_activa = sheets[0]
            for sheet_name in sheets:
                ws = wb[sheet_name]
                if ws.max_row and ws.max_row > 1:
                    hoja_activa = sheet_name
                    break

            df = pd.read_excel(path_archivo, sheet_name=hoja_activa)
            total_filas = len(df)
            columnas = _columnas_desde_dataframe(df)
            wb.close()

        carga.hoja_activa = hoja_activa
        carga.columnas_raw = columnas
        carga.total_filas = total_filas
        carga.save(update_fields=["hoja_activa", "columnas_raw", "total_filas"])

        # Recuperar el avance de mapeo de la última carga de esta fuente,
        # copiándolo a la carga nueva (solo columnas que siguen existiendo).
        mapeos_previos = []
        carga_previa = (
            CargaArchivo.objects.filter(fuente=fuente, mapeos__isnull=False)
            .exclude(pk=carga.pk)
            .order_by("-created_at")
            .first()
        )
        if carga_previa:
            nombres_actuales = {c["nombre"] for c in columnas}
            # Los atributos manuales (constantes) no dependen de las columnas del
            # archivo, así que siempre se conservan.
            copias = [
                m for m in carga_previa.mapeos.all()
                if m.columna_origen in nombres_actuales or m.transformacion == "constante"
            ]
            MapeoColumna.objects.bulk_create([
                MapeoColumna(
                    carga=carga,
                    columna_origen=m.columna_origen,
                    modelo_destino=m.modelo_destino,
                    campo_destino=m.campo_destino,
                    transformacion=m.transformacion,
                    mapeo_valores=m.mapeo_valores,
                    valor_constante=m.valor_constante,
                    estrategia_nulos=m.estrategia_nulos,
                    valor_relleno_manual=m.valor_relleno_manual,
                )
                for m in copias
            ])
            mapeos_previos = [
                {
                    "columna_origen": m.columna_origen,
                    "modelo_destino": m.modelo_destino,
                    "campo_destino": m.campo_destino,
                    "transformacion": m.transformacion,
                    "mapeo_valores": m.mapeo_valores,
                    "valor_constante": m.valor_constante,
                    "estrategia_nulos": m.estrategia_nulos,
                    "valor_relleno_manual": m.valor_relleno_manual,
                }
                for m in copias
            ]

        return JsonResponse({
            "carga_id": carga.pk,
            "sheets": sheets,
            "hoja_activa": hoja_activa,
            "total_filas": total_filas,
            "columnas": columnas,
            "mapeos": mapeos_previos,
        }, json_dumps_params={"ensure_ascii": False})

    except (ValueError, FileNotFoundError) as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


def campos_destino(request):
    modelos = {}
    grupos = {}
    orden_modelo = 0
    for orden, grupo in enumerate(SECCIONES_ETL):
        for nombre in grupo["entidades"]:
            try:
                modelo_cls = apps.get_model("app", nombre)
            except LookupError:
                continue
            campos = []
            for field in modelo_cls._meta.get_fields():
                if field.is_relation and not hasattr(field, "column"):
                    continue
                if field.name in ("id", "created_at", "updated_at"):
                    continue
                campos.append(campo_to_catalogo(field))
            if campos:
                modelos[nombre] = campos
                grupos[nombre] = {
                    "nombre": grupo["nombre"],
                    "icono": grupo["icono"],
                    "orden": orden,
                    "orden_modelo": orden_modelo,
                }
                orden_modelo += 1
    return JsonResponse({"modelos": modelos, "grupos": grupos}, json_dumps_params={"ensure_ascii": False})


@csrf_exempt
def mapeo_carga(request, fuente_id, carga_id):
    try:
        carga = CargaArchivo.objects.get(pk=carga_id, fuente_id=fuente_id)
    except CargaArchivo.DoesNotExist:
        return JsonResponse({"error": "Carga no encontrada"}, status=404)

    if request.method == "GET":
        mapeos = list(
            carga.mapeos.values(
                "columna_origen", "modelo_destino", "campo_destino",
                "transformacion", "mapeo_valores", "valor_constante",
                "estrategia_nulos", "valor_relleno_manual",
            )
        )
        return JsonResponse({
            "carga_id": carga.pk,
            "fuente_id": carga.fuente_id,
            "fuente_nombre": carga.fuente.nombre,
            "estado": carga.estado,
            "columnas_raw": carga.columnas_raw,
            "total_filas": carga.total_filas,
            "mapeos": mapeos,
        }, json_dumps_params={"ensure_ascii": False})

    if request.method == "POST":
        try:
            body = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "JSON inválido"}, status=400)

        items = body.get("mapeos")
        if not isinstance(items, list):
            return JsonResponse({"error": "Se esperaba un array en 'mapeos'"}, status=400)

        # parcial=True: guardado incremental de columnas sueltas; no cambia el estado
        parcial = bool(body.get("parcial"))

        try:
            guardados = 0
            enviados = set()
            for item in items:
                columna_origen = (item.get("columna_origen") or "").strip()
                if not columna_origen:
                    continue
                MapeoColumna.objects.update_or_create(
                    carga=carga,
                    columna_origen=columna_origen,
                    defaults={
                        "modelo_destino": item.get("modelo_destino", ""),
                        "campo_destino": item.get("campo_destino", ""),
                        "transformacion": item.get("transformacion", "directo"),
                        "mapeo_valores": item.get("mapeo_valores") or {},
                        "valor_constante": item.get("valor_constante", ""),
                        "estrategia_nulos": item.get("estrategia_nulos")
                        if item.get("estrategia_nulos") in dict(MapeoColumna.ESTRATEGIA_NULOS_CHOICES)
                        else "dejar_null",
                        "valor_relleno_manual": item.get("valor_relleno_manual", ""),
                    },
                )
                enviados.add(columna_origen)
                guardados += 1

            # El frontend siempre envía el estado completo (columnas + atributos
            # manuales): lo que ya no venga se elimina (p. ej. un atributo manual
            # que se quitó o se re-apuntó a otro campo).
            carga.mapeos.exclude(columna_origen__in=enviados).delete()

            if not parcial:
                carga.estado = "mapeado"
                carga.save(update_fields=["estado"])

            return JsonResponse({"ok": True, "guardados": guardados})
        except Exception as exc:
            return JsonResponse({"error": str(exc)}, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)


def _to_python(val):
    if val is None:
        return None
    if isinstance(val, float) and math.isnan(val):
        return None
    try:
        import numpy as np
        if isinstance(val, (np.integer,)):
            return int(val)
        if isinstance(val, (np.floating,)):
            return None if math.isnan(float(val)) else float(val)
        if isinstance(val, (np.bool_,)):
            return bool(val)
    except ImportError:
        pass
    return val


def _es_vacio(val):
    if val is None:
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    if isinstance(val, str) and val.strip() == "":
        return True
    return False


def _es_fk(field):
    """True para ForeignKey y OneToOneField (esta última no aparece con ese
    nombre de clase, pero se comporta igual como relación N:1 con columna)."""
    return type(field).__name__ in ("ForeignKey", "OneToOneField")


def _choices_de_campo(field):
    """Choices efectivas de un campo: las declaradas en Django, o -para FK sin
    choices propias- las instancias existentes del modelo relacionado."""
    choices = getattr(field, "choices", None)
    if choices:
        return [(str(valor), etiqueta) for valor, etiqueta in choices]
    if _es_fk(field):
        return [(c["valor"], c["etiqueta"]) for c in fk_choices(field)]
    return None


_RE_HORA_MERIDIANO = re.compile(r"^(\d{1,2}):(\d{2})(?::(\d{2}(?:\.\d+)?))?\s*([ap]\.?\s*m\.?)\s*$", re.IGNORECASE)


def _sugerir_hora(valor):
    """Si `valor` tiene forma de hora con sufijo a. m./p. m. pero el formato
    estricto (el que exige Django) lo rechaza -por contradicción entre la
    hora en 24h y el sufijo, p. ej. "14:29:00 a. m."-, devuelve una
    *sugerencia* de interpretación (ignorando el sufijo) para que el usuario
    la confirme o la corrija en la UI. No se aplica automáticamente. Si el
    valor no aplica o ya es válido tal cual, devuelve None."""
    match = _RE_HORA_MERIDIANO.match(str(valor).strip())
    if not match:
        return None
    try:
        pd.to_datetime(valor)
        return None  # ya es válido en formato estricto, no hay ambigüedad
    except Exception:
        pass
    hora, minuto, segundo = match.group(1), match.group(2), match.group(3) or "00"
    candidato = f"{hora}:{minuto}:{segundo}"
    try:
        return pd.to_datetime(candidato).strftime("%H:%M:%S")
    except Exception:
        return None


def _aplicar_estrategia_nulos(df, mapeos):
    """Aplica sobre el propio DataFrame -antes de validar/previsualizar/
    importar- lo que el usuario eligió para los nulos de cada columna
    mapeada: 'rellenar' repite hacia abajo el último valor visto; 'manual'
    usa un valor fijo. 'dejar_null' e 'ignorar_fila' no tocan el DataFrame
    ('ignorar_fila' se resuelve fila por fila en _procesar_filas)."""
    for mapeo in mapeos:
        if mapeo.columna_origen not in df.columns:
            continue
        if mapeo.estrategia_nulos == "rellenar":
            df[mapeo.columna_origen] = df[mapeo.columna_origen].ffill()
        elif mapeo.estrategia_nulos == "manual" and not _es_vacio(mapeo.valor_relleno_manual):
            df[mapeo.columna_origen] = df[mapeo.columna_origen].fillna(mapeo.valor_relleno_manual)


def _resolver_valor_columna(val, mapeo):
    """Aplica la traducción de mapeo_valores (origen -> destino) de una columna
    mapeada a un campo con choices, igual que hace la UI antes de guardar."""
    if _es_vacio(val):
        return None
    if mapeo.mapeo_valores:
        return mapeo.mapeo_valores.get(str(val), val)
    return val


def _validar_columna(df, mapeo):
    columna_origen = mapeo.columna_origen
    modelo_destino = mapeo.modelo_destino
    campo_destino = mapeo.campo_destino
    try:
        modelo_cls = apps.get_model("app", modelo_destino)
    except LookupError:
        return {"advertencia": "modelo_o_campo_no_encontrado"}

    try:
        field = modelo_cls._meta.get_field(campo_destino)
    except Exception:
        return {"advertencia": "modelo_o_campo_no_encontrado"}

    if columna_origen not in df.columns:
        return {"advertencia": "modelo_o_campo_no_encontrado"}

    serie = df[columna_origen]
    errores = []

    campo_requerido = not getattr(field, "blank", True) and not getattr(field, "null", True)
    max_length = getattr(field, "max_length", None)
    choices = _choices_de_campo(field)
    choices_valores = {v for v, _ in choices} if choices else None

    tipo_campo = type(field).__name__

    for idx, val_crudo in serie.items():
        fila = int(idx) + 2
        val = _resolver_valor_columna(val_crudo, mapeo)
        py_val = _to_python(val)

        if val is None:
            # "ignorar_fila": el usuario ya decidió que estas filas no crean
            # el registro, así que no es un error de validación bloqueante.
            if campo_requerido and mapeo.estrategia_nulos != "ignorar_fila":
                errores.append({
                    "fila": fila,
                    "valor": None,
                    "tipo": "nulo_obligatorio",
                    "mensaje": "Campo requerido vacío",
                })
            continue

        if tipo_campo in ("FloatField", "DecimalField"):
            try:
                float(val)
            except (ValueError, TypeError):
                errores.append({
                    "fila": fila,
                    "valor": py_val,
                    "tipo": "tipo_invalido",
                    "mensaje": "No se puede convertir a número",
                })
                continue

        elif tipo_campo in ("IntegerField", "PositiveIntegerField", "PositiveSmallIntegerField",
                            "SmallIntegerField", "BigIntegerField"):
            try:
                int(float(val))
            except (ValueError, TypeError):
                errores.append({
                    "fila": fila,
                    "valor": py_val,
                    "tipo": "tipo_invalido",
                    "mensaje": "No se puede convertir a entero",
                })
                continue

        elif tipo_campo in ("DateField", "DateTimeField"):
            try:
                pd.to_datetime(val)
            except Exception:
                errores.append({
                    "fila": fila,
                    "valor": py_val,
                    "tipo": "tipo_invalido",
                    "mensaje": "No se puede interpretar como fecha",
                })
                continue

        elif tipo_campo == "TimeField":
            try:
                pd.to_datetime(val)
            except Exception:
                errores.append({
                    "fila": fila,
                    "valor": py_val,
                    "tipo": "tipo_invalido",
                    "mensaje": "No se puede interpretar como hora",
                })
                continue

        if max_length is not None:
            if len(str(val)) > max_length:
                errores.append({
                    "fila": fila,
                    "valor": py_val,
                    "tipo": "longitud_excedida",
                    "mensaje": f"Longitud {len(str(val))} supera el máximo de {max_length}",
                })
                continue

        if choices_valores is not None:
            if val not in choices_valores and str(val) not in choices_valores:
                errores.append({
                    "fila": fila,
                    "valor": py_val,
                    "tipo": "fuera_de_vocabulario",
                    "mensaje": f"Valor no permitido. Opciones: {sorted(str(v) for v in choices_valores)}",
                })

    total = len(serie)
    return {
        "columna": columna_origen,
        "modelo_destino": modelo_destino,
        "campo_destino": campo_destino,
        "total": total,
        "ok": total - len(errores),
        "errores": errores,
    }


def _validar_constante(mapeo, total_filas):
    """Valida el valor fijo de un atributo manual contra el campo destino (una sola vez, aplica a todas las filas)."""
    try:
        modelo_cls = apps.get_model("app", mapeo.modelo_destino)
        field = modelo_cls._meta.get_field(mapeo.campo_destino)
    except Exception:
        return {"advertencia": "modelo_o_campo_no_encontrado"}

    val = mapeo.valor_constante
    errores = []
    campo_requerido = not getattr(field, "blank", True) and not getattr(field, "null", True)
    max_length = getattr(field, "max_length", None)
    choices = _choices_de_campo(field)
    tipo_campo = type(field).__name__

    if _es_vacio(val):
        if campo_requerido:
            errores.append({"fila": None, "valor": None, "tipo": "nulo_obligatorio", "mensaje": "Campo requerido vacío"})
    else:
        if tipo_campo in ("FloatField", "DecimalField"):
            try:
                float(val)
            except (ValueError, TypeError):
                errores.append({"fila": None, "valor": val, "tipo": "tipo_invalido", "mensaje": "No se puede convertir a número"})
        elif tipo_campo in ("IntegerField", "PositiveIntegerField", "PositiveSmallIntegerField",
                            "SmallIntegerField", "BigIntegerField"):
            try:
                int(float(val))
            except (ValueError, TypeError):
                errores.append({"fila": None, "valor": val, "tipo": "tipo_invalido", "mensaje": "No se puede convertir a entero"})
        elif tipo_campo in ("DateField", "DateTimeField"):
            try:
                pd.to_datetime(val)
            except Exception:
                errores.append({"fila": None, "valor": val, "tipo": "tipo_invalido", "mensaje": "No se puede interpretar como fecha"})
        elif tipo_campo == "TimeField":
            try:
                pd.to_datetime(val)
            except Exception:
                errores.append({"fila": None, "valor": val, "tipo": "tipo_invalido", "mensaje": "No se puede interpretar como hora"})
        if not errores and max_length is not None and len(str(val)) > max_length:
            errores.append({"fila": None, "valor": val, "tipo": "longitud_excedida", "mensaje": f"Longitud {len(str(val))} supera el máximo de {max_length}"})
        if not errores and choices is not None and str(val) not in {v for v, _ in choices}:
            errores.append({"fila": None, "valor": val, "tipo": "fuera_de_vocabulario", "mensaje": f"Valor no permitido. Opciones: {sorted(v for v, _ in choices)}"})

    return {
        "columna": mapeo.columna_origen,
        "modelo_destino": mapeo.modelo_destino,
        "campo_destino": mapeo.campo_destino,
        "total": total_filas,
        "ok": 0 if errores else total_filas,
        "errores": errores,
    }


def _valores_fila_modelo(df, mapeos_modelo, fila_idx):
    """Valores {campo_destino: valor} que tendría una instancia del modelo en
    una fila dada, según sus mapeos (columna directa o constante)."""
    valores = {}
    for m in mapeos_modelo:
        if m.transformacion == "constante":
            valor = None if _es_vacio(m.valor_constante) else m.valor_constante
        else:
            val_crudo = df[m.columna_origen].iloc[fila_idx]
            valor = _resolver_valor_columna(val_crudo, m)
        valores[m.campo_destino] = valor
    return valores


def _validar_obligatorios_unidad_muestreo(mapeos, modelos_incluidos, total_filas):
    """Reglas de negocio de UnidadMuestreo que no se derivan de blank/null del
    modelo (por eso _validar_columna/_validar_constante no las cubren):
    - 'nombre' es obligatorio para toda unidad de muestreo, así que debe
      quedar mapeado (columna o atributo manual) antes de guardar la sección.
    - 'unidad_experimental' debe quedar resuelto sí o sí: mapeado explícito,
      o heredado automáticamente porque 'UnidadExperimental' se importa en la
      misma tanda (ver el vínculo automático por fila en importar_carga)."""
    if "UnidadMuestreo" not in modelos_incluidos:
        return None

    campos_um = {m.campo_destino for m in mapeos if m.modelo_destino == "UnidadMuestreo"}
    errores = []

    if "nombre" not in campos_um:
        errores.append({
            "fila": None, "valor": None, "tipo": "campo_obligatorio_sin_mapear",
            "mensaje": (
                "El campo 'nombre' de Unidad de Muestreo no está mapeado (ni por "
                "columna ni como atributo manual). Es obligatorio: toda unidad de "
                "muestreo necesita un nombre."
            ),
        })

    if "unidad_experimental" not in campos_um and "UnidadExperimental" not in modelos_incluidos:
        errores.append({
            "fila": None, "valor": None, "tipo": "campo_obligatorio_sin_mapear",
            "mensaje": (
                "Unidad de Muestreo no tiene 'unidad_experimental' mapeada, y la "
                "sección 'Unidad Experimental' no forma parte de esta importación, "
                "así que no se puede vincular automáticamente. Mapea la columna, o "
                "guarda primero la sección Unidad Experimental."
            ),
        })

    return {
        "columna": "Unidad de Muestreo (campos obligatorios)",
        "modelo_destino": "UnidadMuestreo",
        "campo_destino": "",
        "total": total_filas,
        "ok": 0 if errores else total_filas,
        "errores": errores,
    }


def _validar_unicidad_unidad_experimental(carga, df):
    """Unidad Experimental es única por (proyecto, nombre): la fuente debe
    tener un proyecto asociado, y si ese nombre ya existe en el proyecto con
    otros datos (p. ej. otra descripción), se avisa acá en vez de fallar con
    un error de base de datos al intentar crearla."""
    mapeos_ue = [m for m in carga.mapeos.exclude(modelo_destino="") if m.modelo_destino == "UnidadExperimental"]
    mapeo_nombre = next((m for m in mapeos_ue if m.campo_destino == "nombre"), None)
    if mapeo_nombre is None:
        return None

    errores = []
    proyecto = carga.fuente.proyecto
    if proyecto is None:
        errores.append({
            "fila": None, "valor": None, "tipo": "sin_proyecto",
            "mensaje": "La fuente de datos no tiene un proyecto asociado. Asigna un proyecto a la fuente antes de continuar.",
        })
    else:
        UnidadExperimental = apps.get_model("app", "UnidadExperimental")
        vistos = {}
        for fila_idx in range(carga.total_filas):
            valores = _valores_fila_modelo(df, mapeos_ue, fila_idx)
            nombre = valores.get("nombre")
            if _es_vacio(nombre):
                continue
            fila = fila_idx + 2

            previo = vistos.get(nombre)
            if previo is not None and previo != valores:
                errores.append({
                    "fila": fila, "valor": nombre, "tipo": "conflicto_unicidad",
                    "mensaje": f"'{nombre}' aparece con datos distintos en otra fila de este mismo archivo.",
                })
            vistos[nombre] = valores

            existente = UnidadExperimental.objects.filter(proyecto=proyecto, nombre=nombre).first()
            if existente is not None:
                for campo, valor in valores.items():
                    if campo == "nombre" or valor is None:
                        continue
                    if str(getattr(existente, campo)) != str(valor):
                        errores.append({
                            "fila": fila, "valor": nombre, "tipo": "conflicto_unicidad",
                            "mensaje": (
                                f"Ya existe una Unidad Experimental '{nombre}' en el proyecto "
                                f"'{proyecto.nombre}' con datos distintos (campo '{campo}'). "
                                "Cambia el nombre o corrige el valor para que coincida."
                            ),
                        })
                        break

    filas_con_error = {e["fila"] for e in errores if e["fila"] is not None}
    ok = 0 if any(e["tipo"] == "sin_proyecto" for e in errores) else carga.total_filas - len(filas_con_error)
    return {
        "columna": "Unidad Experimental (nombre único por proyecto)",
        "modelo_destino": "UnidadExperimental",
        "campo_destino": "nombre",
        "total": carga.total_filas,
        "ok": ok,
        "errores": errores,
    }


@csrf_exempt
def validar_carga(request, fuente_id, carga_id):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        carga = CargaArchivo.objects.get(pk=carga_id, fuente_id=fuente_id)
    except CargaArchivo.DoesNotExist:
        return JsonResponse({"error": "Carga no encontrada"}, status=404)

    try:
        path = _resolver_ruta_fuente(carga.fuente)
        nombre = str(path).lower()
        if nombre.endswith(".csv"):
            df = _leer_csv(path)
        else:
            df = pd.read_excel(path, sheet_name=carga.hoja_activa)

        mapeos = carga.mapeos.exclude(modelo_destino="")
        _aplicar_estrategia_nulos(df, mapeos)

        resultados = []
        filas_con_error = set()

        for mapeo in mapeos:
            if mapeo.transformacion == "constante":
                resultado = _validar_constante(mapeo, carga.total_filas)
            else:
                resultado = _validar_columna(df, mapeo)
            if "advertencia" not in resultado:
                for e in resultado["errores"]:
                    filas_con_error.add(e["fila"])
            resultados.append(resultado)

        resultado_ue = _validar_unicidad_unidad_experimental(carga, df)
        if resultado_ue is not None:
            for e in resultado_ue["errores"]:
                filas_con_error.add(e["fila"])
            resultados.append(resultado_ue)

        modelos_incluidos = {m.modelo_destino for m in mapeos}
        resultado_um = _validar_obligatorios_unidad_muestreo(mapeos, modelos_incluidos, carga.total_filas)
        if resultado_um is not None:
            for e in resultado_um["errores"]:
                filas_con_error.add(e["fila"])
            resultados.append(resultado_um)

        columnas_con_errores = sum(
            1 for r in resultados
            if "advertencia" not in r and len(r["errores"]) > 0
        )
        total_errores = sum(
            len(r["errores"]) for r in resultados if "advertencia" not in r
        )

        return JsonResponse({
            "resumen": {
                "total_filas": carga.total_filas,
                "columnas_mapeadas": mapeos.count(),
                "columnas_con_errores": columnas_con_errores,
                "total_errores": total_errores,
                "filas_limpias": carga.total_filas - len(filas_con_error),
                "filas_con_errores": len(filas_con_error),
            },
            "columnas": resultados,
        }, json_dumps_params={"ensure_ascii": False})

    except (ValueError, FileNotFoundError) as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


_ORDEN_GRUPO_MODELO = {
    nombre: orden
    for orden, grupo in enumerate(SECCIONES_ETL)
    for nombre in grupo["entidades"]
}


def _grupo_de_modelo(modelo):
    return _ORDEN_GRUPO_MODELO.get(modelo, len(SECCIONES_ETL))


def _orden_topologico(modelos):
    """Ordena `modelos` (nombres) de forma que cada uno quede después de
    cualquier otro modelo de la misma lista al que referencia por FK, para
    poder crearlos en ese orden (p. ej. UnidadMuestreo antes que Parcela)."""
    dependencias = {}
    for modelo in modelos:
        try:
            modelo_cls = apps.get_model("app", modelo)
        except LookupError:
            dependencias[modelo] = set()
            continue
        dependencias[modelo] = {
            field.related_model.__name__
            for field in modelo_cls._meta.get_fields()
            if _es_fk(field) and field.related_model.__name__ in modelos
        }

    ordenados = []
    vistos = set()

    def visitar(modelo):
        if modelo in vistos:
            return
        vistos.add(modelo)
        for dependencia in dependencias.get(modelo, ()):
            visitar(dependencia)
        ordenados.append(modelo)

    for modelo in modelos:
        visitar(modelo)
    return ordenados


def _coercionar_valor(valor, field):
    tipo_campo = type(field).__name__
    if tipo_campo in ("FloatField", "DecimalField"):
        return float(valor)
    if tipo_campo in ("IntegerField", "PositiveIntegerField", "PositiveSmallIntegerField",
                      "SmallIntegerField", "BigIntegerField"):
        return int(float(valor))
    if tipo_campo in ("DateField", "DateTimeField"):
        ts = pd.to_datetime(valor)
        return ts.date() if tipo_campo == "DateField" else ts.to_pydatetime()
    if tipo_campo == "TimeField":
        return pd.to_datetime(valor).time()
    if tipo_campo == "BooleanField":
        if isinstance(valor, bool):
            return valor
        return str(valor).strip().lower() in ("1", "true", "verdadero", "si", "sí", "x")
    return valor


_PREVIEW_MAX_POR_MODELO = 200


def _representar_kwargs(kwargs):
    """Convierte los kwargs de creación de una instancia (que pueden incluir
    otras instancias de modelo como valores de FK) en algo serializable y
    legible para mostrar en la tabla de vista previa."""
    out = {}
    for k, v in kwargs.items():
        if hasattr(v, "pk"):
            out[k] = str(v)
        elif hasattr(v, "isoformat"):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out


def _procesar_filas(df, orden, mapeos_por_modelo, carga, resumen_modelos, capturar_detalle=False):
    """Recorre cada fila del archivo y crea/reutiliza (get_or_create /
    update_or_create) una instancia por modelo en `orden`, vinculando por FK
    las que se crearon para la misma fila. Se usa tanto para importar de
    verdad como, dentro de una transacción que se revierte, para la vista
    previa (capturar_detalle=True) sin escribir nada permanente."""
    detalle = {} if capturar_detalle else None
    # A diferencia de `detalle` (limitado a _PREVIEW_MAX_POR_MODELO para la
    # UI de vista previa), esto guarda TODOS los pk tocados por esta corrida,
    # sin límite, para poder filtrar después "solo lo de esta carga".
    pks_por_modelo = {}

    for fila_idx in range(len(df)):
        instancias_fila = {}
        for modelo in orden:
            modelo_cls = apps.get_model("app", modelo)
            kwargs = {}
            incompleto = False

            for mapeo in mapeos_por_modelo.get(modelo, []):
                field = modelo_cls._meta.get_field(mapeo.campo_destino)

                if mapeo.transformacion == "constante":
                    valor = None if _es_vacio(mapeo.valor_constante) else mapeo.valor_constante
                else:
                    val_crudo = df[mapeo.columna_origen].iloc[fila_idx]
                    valor = _resolver_valor_columna(val_crudo, mapeo)

                if valor is None:
                    # "ignorar_fila": el usuario decidió explícitamente no
                    # crear el registro de este modelo cuando esta columna
                    # viene vacía, sin importar si el campo admite nulos.
                    if mapeo.estrategia_nulos == "ignorar_fila" or not (
                        getattr(field, "blank", True) or getattr(field, "null", True)
                    ):
                        incompleto = True
                    continue

                if _es_fk(field):
                    fk_modelo = field.related_model.__name__
                    if fk_modelo in instancias_fila:
                        kwargs[mapeo.campo_destino] = instancias_fila[fk_modelo]
                    else:
                        try:
                            kwargs[mapeo.campo_destino] = field.related_model.objects.get(pk=valor)
                        except (field.related_model.DoesNotExist, ValueError, TypeError):
                            incompleto = True
                else:
                    try:
                        kwargs[mapeo.campo_destino] = _coercionar_valor(valor, field)
                    except (ValueError, TypeError):
                        incompleto = True

            # Vincular automáticamente los FK hacia otros modelos ya
            # creados en esta misma fila, aunque el usuario no haya
            # mapeado esa columna explícitamente (p. ej. Parcela se
            # asocia solo a la UnidadMuestreo de su misma fila).
            for field in modelo_cls._meta.get_fields():
                if not _es_fk(field) or field.name in kwargs:
                    continue
                fk_modelo = field.related_model.__name__
                if fk_modelo in instancias_fila:
                    kwargs[field.name] = instancias_fila[fk_modelo]

            if incompleto or not kwargs:
                continue

            if modelo == "UnidadMuestreo":
                kwargs.setdefault("fuente_datos", carga.fuente)

            if modelo == "UnidadExperimental":
                kwargs.setdefault("proyecto", carga.fuente.proyecto)

            # Si el modelo tiene un vínculo OneToOne (p. ej. Parcela →
            # unidad_muestreo), ese vínculo es su clave real: solo puede
            # existir una fila por unidad, así que el resto de los
            # campos se actualizan en vez de intentar crear otra fila
            # (que violaría la restricción única).
            campos_clave = {
                nombre: valor
                for nombre, valor in kwargs.items()
                if type(modelo_cls._meta.get_field(nombre)).__name__ == "OneToOneField"
            }
            if campos_clave:
                defaults = {k: v for k, v in kwargs.items() if k not in campos_clave}
                obj, creado = modelo_cls.objects.update_or_create(**campos_clave, defaults=defaults)
            else:
                obj, creado = modelo_cls.objects.get_or_create(**kwargs)

            instancias_fila[modelo] = obj
            resumen_modelos[modelo]["creados" if creado else "reutilizados"] += 1
            pks_por_modelo.setdefault(modelo, set()).add(obj.pk)

            if capturar_detalle:
                bucket = detalle.setdefault(modelo, {})
                if obj.pk not in bucket and len(bucket) < _PREVIEW_MAX_POR_MODELO:
                    bucket[obj.pk] = {
                        "accion": "creado" if creado else "reutilizado",
                        "campos": _representar_kwargs(kwargs),
                    }

    return detalle, pks_por_modelo


def _preparar_importacion(carga, hasta_grupo_solicitado):
    """Validaciones y datos comunes a importar_carga y previsualizar_carga:
    resuelve `hasta_grupo`, arma los mapeos de la sección y valida el archivo.
    Devuelve (mapeos, hasta_grupo, grupo_maximo, error_response) — si hay
    error, todo lo demás es None y `error_response` ya es el JsonResponse a
    devolver tal cual."""
    todos_mapeos = list(carga.mapeos.exclude(modelo_destino=""))
    if not todos_mapeos:
        return None, None, None, JsonResponse({"error": "La carga no tiene mapeos definidos."}, status=400)

    grupo_maximo = max(_grupo_de_modelo(m.modelo_destino) for m in todos_mapeos)
    hasta_grupo = hasta_grupo_solicitado if hasta_grupo_solicitado is not None else grupo_maximo
    try:
        hasta_grupo = int(hasta_grupo)
    except (TypeError, ValueError):
        return None, None, None, JsonResponse({"error": "hasta_grupo inválido"}, status=400)

    mapeos = [m for m in todos_mapeos if _grupo_de_modelo(m.modelo_destino) <= hasta_grupo]
    if not mapeos:
        return None, None, None, JsonResponse({"error": "No hay mapeos para esta sección."}, status=400)

    return mapeos, hasta_grupo, grupo_maximo, None


def _validar_seccion(carga, df, mapeos):
    """Corre todas las validaciones (por columna + reglas de negocio) sobre
    el subconjunto `mapeos` de esta sección. Devuelve (resultados, total_errores,
    filas_con_error)."""
    resultados = []
    filas_con_error = set()
    for mapeo in mapeos:
        if mapeo.transformacion == "constante":
            resultado = _validar_constante(mapeo, carga.total_filas)
        else:
            resultado = _validar_columna(df, mapeo)
        if "advertencia" not in resultado:
            for e in resultado["errores"]:
                filas_con_error.add(e["fila"])
        resultados.append(resultado)

    resultado_ue = _validar_unicidad_unidad_experimental(carga, df)
    if resultado_ue is not None:
        for e in resultado_ue["errores"]:
            filas_con_error.add(e["fila"])
        resultados.append(resultado_ue)

    modelos_incluidos = {m.modelo_destino for m in mapeos}
    resultado_um = _validar_obligatorios_unidad_muestreo(mapeos, modelos_incluidos, carga.total_filas)
    if resultado_um is not None:
        for e in resultado_um["errores"]:
            filas_con_error.add(e["fila"])
        resultados.append(resultado_um)

    total_errores = sum(len(r["errores"]) for r in resultados if "advertencia" not in r)
    return resultados, total_errores, filas_con_error


@csrf_exempt
def previsualizar_carga(request, fuente_id, carga_id):
    """Simula la importación de esta sección (mismo orden de creación/vínculo
    por FK que importar_carga) sin escribir nada permanente en la base, para
    mostrar en el wizard qué se va a crear/reutilizar y con qué referencias
    antes de confirmar el guardado."""
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        carga = CargaArchivo.objects.get(pk=carga_id, fuente_id=fuente_id)
    except CargaArchivo.DoesNotExist:
        return JsonResponse({"error": "Carga no encontrada"}, status=404)

    try:
        body = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)

    mapeos, hasta_grupo, grupo_maximo, error_response = _preparar_importacion(carga, body.get("hasta_grupo"))
    if error_response is not None:
        return error_response

    try:
        path = _resolver_ruta_fuente(carga.fuente)
        nombre = str(path).lower()
        if nombre.endswith(".csv"):
            df = _leer_csv(path)
        else:
            df = pd.read_excel(path, sheet_name=carga.hoja_activa)
        _aplicar_estrategia_nulos(df, mapeos)

        resultados, total_errores, filas_con_error = _validar_seccion(carga, df, mapeos)
        if total_errores > 0:
            return JsonResponse({
                "ok": False,
                "resumen": {
                    "total_filas": carga.total_filas,
                    "columnas_mapeadas": len(mapeos),
                    "total_errores": total_errores,
                    "filas_limpias": carga.total_filas - len(filas_con_error),
                    "filas_con_errores": len(filas_con_error),
                },
                "columnas": resultados,
            }, status=400, json_dumps_params={"ensure_ascii": False})

        modelos_incluidos = sorted({m.modelo_destino for m in mapeos})
        orden = _orden_topologico(modelos_incluidos)

        mapeos_por_modelo = {}
        for mapeo in mapeos:
            mapeos_por_modelo.setdefault(mapeo.modelo_destino, []).append(mapeo)

        resumen_modelos = {m: {"creados": 0, "reutilizados": 0} for m in modelos_incluidos}

        with transaction.atomic():
            sid = transaction.savepoint()
            detalle, _pks = _procesar_filas(df, orden, mapeos_por_modelo, carga, resumen_modelos, capturar_detalle=True)
            transaction.savepoint_rollback(sid)

        modelos_preview = {
            modelo: {
                "registros": list(bucket.values()),
                "total": resumen_modelos[modelo]["creados"] + resumen_modelos[modelo]["reutilizados"],
                "truncado": (resumen_modelos[modelo]["creados"] + resumen_modelos[modelo]["reutilizados"]) > len(bucket),
            }
            for modelo, bucket in (detalle or {}).items()
        }

        return JsonResponse({
            "ok": True,
            "hasta_grupo": hasta_grupo,
            "completo": hasta_grupo >= grupo_maximo,
            "modelos": resumen_modelos,
            "detalle": modelos_preview,
        }, json_dumps_params={"ensure_ascii": False})

    except (ValueError, FileNotFoundError) as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


@csrf_exempt
def importar_carga(request, fuente_id, carga_id):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        carga = CargaArchivo.objects.get(pk=carga_id, fuente_id=fuente_id)
    except CargaArchivo.DoesNotExist:
        return JsonResponse({"error": "Carga no encontrada"}, status=404)

    try:
        body = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)

    mapeos, hasta_grupo, grupo_maximo, error_response = _preparar_importacion(carga, body.get("hasta_grupo"))
    if error_response is not None:
        return error_response

    try:
        path = _resolver_ruta_fuente(carga.fuente)
        nombre = str(path).lower()
        if nombre.endswith(".csv"):
            df = _leer_csv(path)
        else:
            df = pd.read_excel(path, sheet_name=carga.hoja_activa)
        _aplicar_estrategia_nulos(df, mapeos)

        # 1) Validar solo el subconjunto de mapeos de esta sección; no se
        # escribe nada en la base si queda algún error.
        resultados, total_errores, filas_con_error = _validar_seccion(carga, df, mapeos)
        if total_errores > 0:
            return JsonResponse({
                "ok": False,
                "resumen": {
                    "total_filas": carga.total_filas,
                    "columnas_mapeadas": len(mapeos),
                    "total_errores": total_errores,
                    "filas_limpias": carga.total_filas - len(filas_con_error),
                    "filas_con_errores": len(filas_con_error),
                },
                "columnas": resultados,
            }, status=400, json_dumps_params={"ensure_ascii": False})

        # 2) Importar: crear/reutilizar instancias, en orden de dependencia FK.
        modelos_incluidos = sorted({m.modelo_destino for m in mapeos})
        orden = _orden_topologico(modelos_incluidos)

        mapeos_por_modelo = {}
        for mapeo in mapeos:
            mapeos_por_modelo.setdefault(mapeo.modelo_destino, []).append(mapeo)

        resumen_modelos = {m: {"creados": 0, "reutilizados": 0} for m in modelos_incluidos}

        with transaction.atomic():
            _, pks_por_modelo = _procesar_filas(df, orden, mapeos_por_modelo, carga, resumen_modelos)

            # Acumula (no reemplaza) los pk tocados por esta sección con los de
            # secciones anteriores de la misma carga, para poder mostrar luego
            # "solo lo de esta carga" en el panel de visualización.
            acumulado = carga.pks_importados or {}
            for modelo, pks in pks_por_modelo.items():
                existentes = set(acumulado.get(modelo, []))
                acumulado[modelo] = sorted(existentes | pks)
            carga.pks_importados = acumulado

            campos_actualizar = ["pks_importados"]
            if hasta_grupo >= grupo_maximo:
                carga.estado = "importado"
                campos_actualizar.append("estado")
            carga.save(update_fields=campos_actualizar)

            if hasta_grupo >= grupo_maximo and carga.fuente.estado != "completo":
                carga.fuente.estado = "completo"
                carga.fuente.save(update_fields=["estado"])

        return JsonResponse({
            "ok": True,
            "hasta_grupo": hasta_grupo,
            "completo": hasta_grupo >= grupo_maximo,
            "modelos": resumen_modelos,
        }, json_dumps_params={"ensure_ascii": False})

    except (ValueError, FileNotFoundError) as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


# ── Panel de visualización: tabla desnormalizada de lo importado ───────────
# Cada "vista" define, para un modelo base, la cadena de FKs hacia arriba que
# se aplana en columnas prefijadas por modelo. Cada tupla de la cadena es
# (nombre_modelo, ruta_de_atributos_desde_el_modelo_base).
#
# "submuestra_co2" llega a UnidadMuestreo/UnidadExperimental/Sitio por el
# vínculo directo MuestraCO2.unidad_muestreo, sin pasar por Anillo: el ETL
# todavía no crea/vincula Anillo (Camara depende de Equipo, cuya sección
# "Torre EC y Flujos" está oculta del wizard), y la relación anillo →
# unidad_muestreo queda pendiente de revisar a futuro (ver app/models/co2.py).
# "unidad_muestreo" es una vista aparte para cargas que todavía no tienen
# MuestraCO2/SubmuestraCO2 importadas (solo unidad de muestreo/experimental).
_VISTAS_DESNORMALIZADAS = {
    "submuestra_co2": {
        "modelo_base": "SubmuestraCO2",
        "orden": ["muestra__fecha", "muestra_id", "n_toma"],
        "cadena": [
            ("SubmuestraCO2", []),
            ("UnidadMedida", ["unidad_medida"]),
            ("MuestraCO2", ["muestra"]),
            ("MuestraAmbiental", ["muestra", "muestra_ambiental"]),
            ("Camara", ["muestra", "camara"]),
            ("Equipo", ["muestra", "camara", "equipo"]),
            ("UnidadMuestreo", ["muestra", "unidad_muestreo"]),
            ("UnidadExperimental", ["muestra", "unidad_muestreo", "unidad_experimental"]),
            ("Sitio", ["muestra", "unidad_muestreo", "sitio"]),
        ],
    },
    "unidad_muestreo": {
        "modelo_base": "UnidadMuestreo",
        "orden": ["unidad_experimental__nombre", "nombre"],
        "cadena": [
            ("UnidadMuestreo", []),
            ("UnidadExperimental", ["unidad_experimental"]),
            ("Sitio", ["sitio"]),
        ],
    },
}


def _select_related_de_cadena(cadena):
    return ["__".join(ruta) for _modelo, ruta in cadena if ruta]


def _campos_planos(modelo_cls):
    """Campos propios de un modelo (sin FKs/relaciones ni id/timestamps),
    listos para aplanar como columnas de la tabla desnormalizada."""
    return [
        f for f in modelo_cls._meta.get_fields()
        if hasattr(f, "column") and not f.is_relation and f.name not in ("id", "created_at", "updated_at")
    ]


def _resolver_ruta(obj, ruta):
    for attr in ruta:
        if obj is None:
            return None
        obj = getattr(obj, attr, None)
    return obj


def _valor_campo_plano(obj, field):
    if obj is None:
        return None
    valor = getattr(obj, field.name, None)
    if valor is None:
        return None
    if getattr(field, "choices", None):
        display = getattr(obj, f"get_{field.name}_display", None)
        if callable(display):
            return display()
    return valor


def datos_carga(request, fuente_id, carga_id):
    try:
        carga = CargaArchivo.objects.get(pk=carga_id, fuente_id=fuente_id)
    except CargaArchivo.DoesNotExist:
        return JsonResponse({"error": "Carga no encontrada"}, status=404)

    nombre_vista = request.GET.get("vista", "submuestra_co2")
    vista = _VISTAS_DESNORMALIZADAS.get(nombre_vista)
    if vista is None:
        return JsonResponse({"error": f"Vista desconocida: {nombre_vista}"}, status=400)

    cadena = vista["cadena"]
    modelo_base = vista["modelo_base"]

    pks = (carga.pks_importados or {}).get(modelo_base, [])
    if not pks:
        return JsonResponse({
            "total": 0, "columnas": [], "filas": [], "vistas": list(_VISTAS_DESNORMALIZADAS.keys()),
            "advertencia": f"Esta carga todavía no tiene {modelo_base} importado.",
        }, json_dumps_params={"ensure_ascii": False})

    try:
        offset = max(int(request.GET.get("offset", 0)), 0)
    except ValueError:
        offset = 0
    try:
        limite = min(max(int(request.GET.get("limite", 500)), 1), 2000)
    except ValueError:
        limite = 500

    columnas = [
        {"clave": f"{modelo_nombre}.{f.name}", "modelo": modelo_nombre, "campo": f.name,
         "verbose_name": str(f.verbose_name)}
        for modelo_nombre, _ruta in cadena
        for f in _campos_planos(apps.get_model("app", modelo_nombre))
    ]
    # Mapa clave ("Modelo.campo") -> ruta ORM, para poder traducir los filtros
    # del usuario a filter(**{"ruta__campo__icontains": ...}).
    ruta_orm_por_clave = {}
    for modelo_nombre, ruta in cadena:
        for f in _campos_planos(apps.get_model("app", modelo_nombre)):
            ruta_orm_por_clave[f"{modelo_nombre}.{f.name}"] = ruta + [f.name]

    ModeloBase = apps.get_model("app", modelo_base)
    qs = ModeloBase.objects.filter(pk__in=pks).select_related(*_select_related_de_cadena(cadena))

    try:
        filtros = json.loads(request.GET.get("filtros") or "{}")
    except json.JSONDecodeError:
        filtros = {}
    for clave, texto in filtros.items():
        ruta_orm = ruta_orm_por_clave.get(clave)
        if not ruta_orm or not str(texto).strip():
            continue
        qs = qs.filter(**{f"{'__'.join(ruta_orm)}__icontains": texto})

    qs = qs.order_by(*vista["orden"])
    total = qs.count()

    filas = []
    for obj in qs[offset:offset + limite]:
        fila = {}
        for modelo_nombre, ruta in cadena:
            related_obj = obj if not ruta else _resolver_ruta(obj, ruta)
            for f in _campos_planos(apps.get_model("app", modelo_nombre)):
                fila[f"{modelo_nombre}.{f.name}"] = _valor_campo_plano(related_obj, f)
        filas.append(fila)

    return JsonResponse({
        "total": total,
        "offset": offset,
        "limite": limite,
        "columnas": columnas,
        "filas": filas,
        "vistas": list(_VISTAS_DESNORMALIZADAS.keys()),
    }, json_dumps_params={"ensure_ascii": False})
